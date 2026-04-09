from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Sum, Q, F
from .models import Cliente, Compra, Pagamento
from .forms import ClienteForm, CompraForm, PagamentoForm
import csv
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def dashboard(request):
    total_clientes = Cliente.objects.count()

    nome = request.GET.get('nome', '')
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')

    clientes_qs = Cliente.objects.all().order_by('nome')

    if nome:
        clientes_qs = clientes_qs.filter(nome__icontains=nome)

    total_a_receber = 0

    for cliente in clientes_qs:
        total_compras = cliente.compras.aggregate(total=Sum('valor'))['total'] or 0
        total_pagamentos = cliente.pagamentos.aggregate(total=Sum('valor'))['total'] or 0

        cliente.saldo_devedor_anno = total_compras - total_pagamentos
        total_a_receber += cliente.saldo_devedor_anno

    paginator = Paginator(clientes_qs, 5)
    clientes_page = paginator.get_page(request.GET.get('page'))

    ultimas_compras = Compra.objects.select_related('cliente').order_by('-data')[:5]

    context = {
        'total_clientes': total_clientes,
        'total_a_receber': total_a_receber,
        'clientes': clientes_page,
        'ultimas_compras': ultimas_compras,
        'nome': nome,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
    }
    return render(request, 'vendas/dashboard.html', context)


def lista_clientes(request):
    nome_filtro = request.GET.get('nome', '').strip()

    clientes_qs = Cliente.objects.all().order_by('nome')

    if nome_filtro:
        clientes_qs = clientes_qs.filter(nome__icontains=nome_filtro)

    for cliente in clientes_qs:
        total_compras = cliente.compras.aggregate(total=Sum('valor'))['total'] or 0
        total_pagamentos = cliente.pagamentos.aggregate(total=Sum('valor'))['total'] or 0
        cliente.saldo_devedor_anno = total_compras - total_pagamentos

    paginator = Paginator(clientes_qs, 10)
    clientes_page = paginator.get_page(request.GET.get('page'))

    return render(request, 'vendas/clientes.html', {
        'clientes': clientes_page,
        'nome_filtro': nome_filtro,
    })


def historico(request):
    nome = request.GET.get('nome', '')
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')

    clientes_qs = Cliente.objects.all().order_by('nome')

    if nome:
        clientes_qs = clientes_qs.filter(nome__icontains=nome)

    clientes_filtrados = []

    for cliente in clientes_qs:
        compras = cliente.compras.all()
        pagamentos = cliente.pagamentos.all()

        if data_inicio and data_fim:
            compras = compras.filter(data__range=[data_inicio, data_fim])
            pagamentos = pagamentos.filter(data__range=[data_inicio, data_fim])

        total_compras = compras.aggregate(total=Sum('valor'))['total'] or 0
        total_pagamentos = pagamentos.aggregate(total=Sum('valor'))['total'] or 0

        cliente.total_compras_anno = total_compras
        cliente.total_pagamentos_anno = total_pagamentos
        cliente.saldo_devedor_anno = total_compras - total_pagamentos

        # só adiciona se tiver movimentação no período
        if total_compras > 0 or total_pagamentos > 0 or not (data_inicio and data_fim):
            clientes_filtrados.append(cliente)

    paginator = Paginator(clientes_filtrados, 10)
    clientes_page = paginator.get_page(request.GET.get('page'))

    return render(request, 'vendas/historico.html', {
        'clientes': clientes_page,
        'nome': nome,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
    })


# Cadastros simples
def criar_cliente(request):
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, '✅ Cliente cadastrado!')
            return redirect('lista_clientes')
    else:
        form = ClienteForm()
    return render(request, 'vendas/cliente_form.html', {'form': form, 'titulo': 'Novo Cliente'})


def nova_compra(request):
    if request.method == 'POST':
        form = CompraForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, '✅ Compra registrada!')
            return redirect('dashboard')
    else:
        form = CompraForm()
    return render(request, 'vendas/compra_form.html', {'form': form, 'titulo': 'Nova Compra'})


def novo_pagamento(request):
    if request.method == 'POST':
        form = PagamentoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, '✅ Pagamento registrado!')
            return redirect('dashboard')
    else:
        form = PagamentoForm()
    return render(request, 'vendas/pagamento_form.html', {'form': form, 'titulo': 'Novo Pagamento'})


# Edição e Exclusão (simples)
def editar_cliente(request, pk):
    obj = get_object_or_404(Cliente, pk=pk)
    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, '✅ Alterado!')
            return redirect('lista_clientes')
    else:
        form = ClienteForm(instance=obj)
    return render(request, 'vendas/cliente_form.html', {'form': form, 'titulo': 'Editar Cliente'})


def editar_compra(request, pk):
    obj = get_object_or_404(Compra, pk=pk)
    if request.method == 'POST':
        form = CompraForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, '✅ Compra alterada!')
            return redirect('historico')
    else:
        form = CompraForm(instance=obj)
    return render(request, 'vendas/compra_form.html', {'form': form, 'titulo': 'Editar Compra'})


def editar_pagamento(request, pk):
    obj = get_object_or_404(Pagamento, pk=pk)
    if request.method == 'POST':
        form = PagamentoForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, '✅ Pagamento alterado!')
            return redirect('historico')
    else:
        form = PagamentoForm(instance=obj)
    return render(request, 'vendas/pagamento_form.html', {'form': form, 'titulo': 'Editar Pagamento'})


def excluir_cliente(request, pk):
    get_object_or_404(Cliente, pk=pk).delete()
    messages.success(request, '🗑️ Cliente excluído!')
    return redirect('lista_clientes')


def excluir_compra(request, pk):
    get_object_or_404(Compra, pk=pk).delete()
    messages.success(request, '🗑️ Compra excluída!')
    return redirect('historico')


def excluir_pagamento(request, pk):
    get_object_or_404(Pagamento, pk=pk).delete()
    messages.success(request, '🗑️ Pagamento excluído!')
    return redirect('historico')


def export_historico_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="historico_completo.csv"'

    writer = csv.writer(response, delimiter=';')  
    writer.writerow(['Cliente', 'Tipo', 'Data', 'Descrição', 'Valor (R$)'])

    clientes = Cliente.objects.prefetch_related('compras', 'pagamentos').order_by('nome')

    for cliente in clientes:
        # Compras
        for compra in cliente.compras.all():
            writer.writerow([
                cliente.nome,
                'COMPRA',
                compra.data.strftime('%d/%m/%Y'),
                compra.descricao_produto,
                f"{compra.valor:.2f}".replace('.', ',')
            ])
        
        # Pagamentos
        for pag in cliente.pagamentos.all():
            writer.writerow([
                cliente.nome,
                'PAGAMENTO',
                pag.data.strftime('%d/%m/%Y'),
                'Pagamento registrado',
                f"{pag.valor:.2f}".replace('.', ',')
            ])

    return response


# ====================== EXPORTAR HISTÓRICO EXCEL ======================
def export_historico_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Histórico Completo"

    # Cabeçalho
    headers = ['Cliente', 'Tipo', 'Data', 'Descrição', 'Valor (R$)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    row = 2
    clientes = Cliente.objects.prefetch_related('compras', 'pagamentos').order_by('nome')

    for cliente in clientes:
        for compra in cliente.compras.all():
            ws.cell(row=row, column=1, value=cliente.nome)
            ws.cell(row=row, column=2, value="COMPRA")
            ws.cell(row=row, column=3, value=compra.data.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=4, value=compra.descricao_produto)
            ws.cell(row=row, column=5, value=float(compra.valor))
            row += 1

        for pag in cliente.pagamentos.all():
            ws.cell(row=row, column=1, value=cliente.nome)
            ws.cell(row=row, column=2, value="PAGAMENTO")
            ws.cell(row=row, column=3, value=pag.data.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=4, value="Pagamento registrado")
            ws.cell(row=row, column=5, value=float(pag.valor))
            row += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="historico_completo.xlsx"'
    wb.save(response)

    return response